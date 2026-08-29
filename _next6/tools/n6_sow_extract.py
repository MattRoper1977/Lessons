#!/usr/bin/env python3
"""N6-F §F2 — flatten the three operative SoW workbooks into one reference.

Operative instruments, per the order and `_passsb/inputs/README.md`:
  BUILD   `_passsb/inputs/Build SOW 2026-2027.xlsx`   (vB — vA is superseded)
  GROW    `_passsg/inputs/GROW SOW 2026-27.xlsx`
  LAUNCH  `_passsl/inputs/LAUNCH KS4 - 2026-27.xlsx`

The weekly sheets carry a strand name only on the first row of each strand
block and leave it blank underneath, so the strand is forward-filled. Every row
keeps its sheet name and 1-based row number, because that pair is how the packs
cite the SoW ('BUILD Weekly - Autumn'!B166) and how a verdict is checked.
"""
import json, os, sys, hashlib
import openpyxl

BOOKS = {
    'BUILD':  '_passsb/inputs/Build SOW 2026-2027.xlsx',
    'GROW':   '_passsg/inputs/GROW SOW 2026-27.xlsx',
    'LAUNCH': '_passsl/inputs/LAUNCH KS4 - 2026-27.xlsx',
}
TERMS = ('Autumn', 'Spring', 'Summer')


def cell(ws, r, c):
    v = ws.cell(r, c).value
    return '' if v is None else str(v).strip()


def extract(lane, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {'lane': lane, 'workbook': path,
           'sha256': hashlib.sha256(open(path, 'rb').read()).hexdigest(),
           'weekly': [], 'grid': [], 'ladder': [], 'qualmap': []}
    for term in TERMS:
        sh = '%s Weekly - %s' % (lane, term)
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        strand = ''
        for r in range(4, ws.max_row + 1):
            s = cell(ws, r, 1)
            if s:
                strand = s
            wk = cell(ws, r, 2)
            outc = cell(ws, r, 3)
            if not wk and not outc:
                continue
            out['weekly'].append({
                'sheet': sh, 'row': r, 'term': term, 'strand': strand,
                'week': wk, 'outcome': outc,
                'programme': cell(ws, r, 4), 'texts': cell(ws, r, 5),
                'accreditation': cell(ws, r, 6),
            })
        gs = '%s - %s' % (lane, term)
        if gs in wb.sheetnames:
            g = wb[gs]
            strand = ''
            for r in range(4, g.max_row + 1):
                s = cell(g, r, 1)
                if s:
                    strand = s
                if not any(cell(g, r, c) for c in range(1, 12)):
                    continue
                out['grid'].append({
                    'sheet': gs, 'row': r, 'term': term, 'strand': strand,
                    'weeks': cell(g, r, 2), 'theme': cell(g, r, 3),
                    'unit': cell(g, r, 4), 'outcomes': cell(g, r, 5),
                    'activities': cell(g, r, 6), 'vocab': cell(g, r, 7),
                    'texts': cell(g, r, 8), 'qualification': cell(g, r, 9),
                    'assessment': cell(g, r, 10), 'send_smsc': cell(g, r, 11),
                })
    if 'Pathway Ladder' in wb.sheetnames:
        ws = wb['Pathway Ladder']
        for r in range(1, ws.max_row + 1):
            row = [cell(ws, r, c) for c in range(1, 5)]
            if any(row):
                out['ladder'].append({'row': r, 'cells': row})
    if 'Qualification Map' in wb.sheetnames:
        ws = wb['Qualification Map']
        for r in range(1, ws.max_row + 1):
            row = [cell(ws, r, c) for c in range(1, 7)]
            if any(row):
                out['qualmap'].append({'row': r, 'cells': row})
    return out


if __name__ == '__main__':
    outdir = sys.argv[1] if len(sys.argv) > 1 else '_next6/sow'
    os.makedirs(outdir, exist_ok=True)
    for lane, path in BOOKS.items():
        d = extract(lane, path)
        json.dump(d, open(os.path.join(outdir, '%s.json' % lane), 'w'), indent=1)
        strands = sorted({w['strand'] for w in d['weekly']})
        wks = sorted({w['week'] for w in d['weekly']})
        print('%-7s %-64s' % (lane, os.path.basename(path)))
        print('        sha256 %s' % d['sha256'])
        print('        %d weekly rows · %d strands · %d distinct week labels'
              % (len(d['weekly']), len(strands), len(wks)))
        for term in TERMS:
            tw = [w for w in d['weekly'] if w['term'] == term]
            ts = sorted({w['strand'] for w in tw})
            per = {}
            for w in tw:
                per.setdefault(w['week'], 0)
                per[w['week']] += 1
            halves = {}
            for w in tw:
                h = w['week'].split('·')[0].strip() if '·' in w['week'] else w['week']
                halves.setdefault(h, set()).add(w['week'])
            print('        %-7s %2d strands  half-terms: %s'
                  % (term, len(ts),
                     ', '.join('%s=%dwk' % (h, len(v)) for h, v in sorted(halves.items()))))
