#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ORDER N6-Z · Z1 — extract the SoW weekly grids, deterministically.

Reads the three operative workbooks and emits every (lane, term, strand, week)
cell with its outcome text. Column layout, confirmed against all three:

  A  Subject / strand      (merged down the strand's block)
  B  HT · Wk               e.g. "Aut1·W1"  (separator is U+00B7)
  C  Weekly learning outcome & key activities
  D  Programme & awarding-body SoW alignment
  E  Age-appropriate texts / resources
  F  Accreditation & assessment evidence

Merged column A is the trap: only the first row of each strand block carries the
strand name, so the reader has to carry it down. Reading column A per-row without
that gives one named strand and six blanks per block.
"""
import hashlib
import json
import os
import re
import sys

import openpyxl

BOOKS = [
    ('BUILD',  '_passsb/inputs/Build SOW 2026-2027.xlsx'),
    ('GROW',   '_passsg/inputs/GROW SOW 2026-27.xlsx'),
    ('LAUNCH', '_passsl/inputs/LAUNCH KS4 - 2026-27.xlsx'),
]
TERMS = ['Autumn', 'Spring', 'Summer']
WK = re.compile(r'^(Aut|Spr|Sum)(\d)\s*[·.\-]\s*W(\d+)', re.I)


def sha(p):
    return hashlib.sha256(open(p, 'rb').read()).hexdigest()


def grid(lane, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for term in TERMS:
        name = '%s Weekly - %s' % (lane, term)
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        strand = None
        for r in range(4, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and str(a).strip():
                strand = str(a).strip()
            b = ws.cell(r, 2).value
            if not b:
                continue
            m = WK.match(str(b).strip())
            if not m:
                continue
            rows.append({
                'lane': lane, 'term': term, 'sheet': name, 'row': r,
                'strand': strand,
                'ht': '%s%s' % (m.group(1).title(), m.group(2)),
                'week': int(m.group(3)),
                'htwk': str(b).strip(),
                'outcome': (str(ws.cell(r, 3).value or '')).strip(),
                'programme': (str(ws.cell(r, 4).value or '')).strip(),
                'texts': (str(ws.cell(r, 5).value or '')).strip(),
                'accred': (str(ws.cell(r, 6).value or '')).strip(),
            })
    return rows


if __name__ == '__main__':
    out = {'books': {}, 'rows': []}
    for lane, path in BOOKS:
        out['books'][lane] = {'path': path, 'sha256': sha(path)}
        out['rows'] += grid(lane, path)
    # Shape report — the order supplies these as a prediction map, so print what
    # is actually there rather than confirming what was predicted.
    import collections
    print('%-8s %-8s %8s %8s  %s' % ('lane', 'term', 'strands', 'rows', 'weeks per half-term'))
    for lane, _ in BOOKS:
        for term in TERMS:
            rs = [r for r in out['rows'] if r['lane'] == lane and r['term'] == term]
            if not rs:
                continue
            strands = sorted({r['strand'] for r in rs})
            per = collections.defaultdict(set)
            for r in rs:
                per[r['ht']].add(r['week'])
            shape = ' '.join('%s=%d(W%d-%d)' % (h, len(w), min(w), max(w))
                             for h, w in sorted(per.items()))
            print('%-8s %-8s %8d %8d  %s' % (lane, term, len(strands), len(rs), shape))
    if len(sys.argv) > 1:
        json.dump(out, open(sys.argv[1], 'w'), indent=1)
        print('\n-> %s  (%d rows)' % (sys.argv[1], len(out['rows'])))
    for lane, v in out['books'].items():
        print('%-7s %s  %s' % (lane, v['sha256'][:16], v['path']))
