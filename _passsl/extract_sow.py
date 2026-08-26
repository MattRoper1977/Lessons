#!/usr/bin/env python3
"""Pass SL — extract the LAUNCH SoW workbook into structured JSON keyed by
(strand, half_term, week). read_only; does not trust max_row.
Produces _passsl/sow_extract/{sow_overview.json, sow_weekly.json, sow_strands.json}."""
import openpyxl, json, re, os

WB = "_passsl/inputs/LAUNCH KS4 - 2026-27.xlsx"
OUT = "_passsl/sow_extract"
os.makedirs(OUT, exist_ok=True)

TERMS = ["Autumn", "Spring", "Summer"]

def norm(c):
    return "" if c is None else str(c).strip()

wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)

# ---------- Overview sheets: LAUNCH - <Term> ----------
# header row (row 3) columns:
OV_COLS = ["strand","htweeks","theme","topic","los","activities","vocab",
           "texts","qual","assessment","send"]
overview = []
for term in TERMS:
    ws = wb[f"LAUNCH - {term}"]
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    cur_strand = None
    for r in rows[3:]:            # skip title + blank + header (rows 1-3)
        if not any(r): continue
        # pad
        r = (r + [""]*11)[:11]
        strand = r[0]
        if strand:
            cur_strand = strand
        rec = {"term": term, "strand": cur_strand}
        for k, v in zip(OV_COLS[1:], r[1:]):
            rec[k] = v
        # parse half-term label from htweeks e.g. "Aut 1 · Sep to Oct (7 wks)"
        m = re.search(r'(Aut|Spr|Sum)\s*([12])', r[1])
        rec["half_term"] = f"{m.group(1)}{m.group(2)}" if m else r[1]
        overview.append(rec)
    print(f"overview {term}: rows kept {sum(1 for x in overview if x['term']==term)}")

# ---------- Weekly sheets: LAUNCH Weekly - <Term> ----------
# header row (row 3): strand | HT·Wk | outcome | SoW alignment | texts | accreditation
weekly = []
for term in TERMS:
    ws = wb[f"LAUNCH Weekly - {term}"]
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    cur_strand = None
    cur_align = cur_texts = cur_accr = ""
    for r in rows[3:]:
        if not any(r): continue
        r = (r + [""]*6)[:6]
        strand, htwk, outcome, align, texts, accr = r
        if strand:
            cur_strand = strand
            cur_align, cur_texts, cur_accr = align, texts, accr
        if not htwk:
            continue
        m = re.search(r'(Aut|Spr|Sum)\s*([12])\s*[·.]?\s*W?(\d+)', htwk)
        if not m:
            continue
        ht = f"{m.group(1)}{m.group(2)}"
        wk = int(m.group(3))
        weekly.append({
            "term": term, "strand": cur_strand, "half_term": ht, "week": wk,
            "outcome": outcome,
            "sow_alignment": align or cur_align,
            "texts": texts or cur_texts,
            "accreditation": accr or cur_accr,
        })
    print(f"weekly {term}: records {sum(1 for x in weekly if x['term']==term)}")

# ---------- strand index (distinct strands, from overview) ----------
seen, strands = set(), []
for rec in overview:
    s = rec["strand"]
    if s and s not in seen:
        seen.add(s); strands.append(s)

wb.close()

json.dump(overview, open(f"{OUT}/sow_overview.json","w"), indent=2, ensure_ascii=False)
json.dump(weekly,   open(f"{OUT}/sow_weekly.json","w"),   indent=2, ensure_ascii=False)
json.dump(strands,  open(f"{OUT}/sow_strands.json","w"),  indent=2, ensure_ascii=False)

print(f"\nDISTINCT STRANDS: {len(strands)}")
for i,s in enumerate(strands,1): print(f"  {i:2d}. {s[:70]}")
# integrity: weeks per (term,strand,half_term)
from collections import Counter
cnt = Counter((w["term"],w["strand"][:30],w["half_term"]) for w in weekly)
bad = {k:v for k,v in cnt.items() if v!=7}
print(f"\n(term,strand,ht) blocks NOT == 7 weeks: {len(bad)}")
for k,v in list(bad.items())[:30]: print(f"   {v}  {k}")
print(f"\nTOTAL overview recs: {len(overview)}  weekly recs: {len(weekly)}")
