import openpyxl, sys, os, json
for name,fp in [("BUILD","_sca1/inputs/BUILD_SOW.xlsx"),("GROW","_sca1/inputs/GROW_SOW.xlsx"),("LAUNCH","_sca1/inputs/LAUNCH_KS4.xlsx")]:
    wb=openpyxl.load_workbook(fp, data_only=True)
    out=[f"### {name}  ({fp})", "SHEETS: "+", ".join(wb.sheetnames)]
    for ws in wb.worksheets:
        out.append(f"\n===== SHEET: {ws.title}  dims={ws.dimensions} =====")
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and str(c.value).strip():
                    v=str(c.value).replace("\n"," ⏎ ").strip()
                    out.append(f"{c.coordinate}\t{v}")
    open(f"_sca1/inputs/{name}_SOW.txt","w").write("\n".join(out))
    print(f"{name}: sheets={len(wb.sheetnames)} lines={len(out)} -> _sca1/inputs/{name}_SOW.txt")
