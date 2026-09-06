#!/usr/bin/env python3
"""SCA-1 CLOSE v2 3b — relabel the Aut1 W1/W2 science rows as baseline weeks.

Owner string (verbatim, not paraphrased):
    Baseline assessment (PythonAnywhere) - no new science content; unit starts W3
(the dash is an em dash in the applied text).

Scope: the per-week 'Weekly - Autumn' sheet of each SoW workbook, rows whose
HT-Wk cell is Aut1-W1 / Aut1-W2 inside a SCIENCE strand block. Column C is the
only per-week content/LO cell (D/E/F are merged across the whole strand block, so
they are strand-level, not week-level, and are left untouched - which also
preserves every merged range).

Verification: every cell of every sheet is dumped before and after; the diff must
contain exactly the intended cells. Merged ranges are compared as sets.
Usage: sow_w1w2.py [--apply]
"""
import sys, json
import openpyxl

LABEL = "Baseline assessment (PythonAnywhere) — no new science content; unit starts W3"
TARGETS = [
    ("_passsb/inputs/Build SOW 2026-2027.xlsx", "BUILD Weekly - Autumn"),
    ("_passsg/inputs/GROW SOW 2026-27.xlsx", "GROW Weekly - Autumn"),
    ("_passsl/inputs/LAUNCH KS4 - 2026-27.xlsx", "LAUNCH Weekly - Autumn"),
]
WEEKS = ("Aut1·W1", "Aut1·W2")


def dump(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells[c.coordinate] = str(c.value)
        out[sn] = {"cells": cells, "merged": sorted(str(r) for r in ws.merged_cells.ranges),
                   "dims": ws.dimensions}
    return out


def find_rows(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet]
    hits, strand = [], None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a:
            strand = str(a)
        if ws.cell(r, 2).value in WEEKS and strand and "cience" in strand:
            hits.append((r, strand, ws.cell(r, 3).value))
    return hits


apply = "--apply" in sys.argv
report = []
for path, sheet in TARGETS:
    before = dump(path)
    hits = find_rows(path, sheet)
    print(f"\n=== {path} :: {sheet}")
    for r, strand, old in hits:
        print(f"  row {r:4d}  strand={strand[:52]!r}")
        print(f"      C{r} OLD: {old!r}")
        print(f"      C{r} NEW: {LABEL!r}")
        report.append({"file": path, "sheet": sheet, "row": r, "cell": f"C{r}",
                       "strand": strand, "old": old, "new": LABEL})
    if apply:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb[sheet]
        for r, _, _ in hits:
            ws.cell(r, 3).value = LABEL
        wb.save(path)
        after = dump(path)
        # prove only the intended cells changed
        diffs = []
        for sn in set(list(before) + list(after)):
            b, a = before.get(sn, {}), after.get(sn, {})
            if b.get("merged") != a.get("merged"):
                diffs.append(f"MERGED RANGES CHANGED in {sn}")
            bc, ac = b.get("cells", {}), a.get("cells", {})
            for k in set(list(bc) + list(ac)):
                if bc.get(k) != ac.get(k):
                    diffs.append(f"{sn}!{k}")
        want = {f"{sheet}!C{r}" for r, _, _ in hits}
        got = set(diffs)
        print(f"  cells changed: {sorted(got)}")
        print(f"  expected     : {sorted(want)}")
        print("  VERDICT:", "OK - only the intended cells changed" if got == want else "MISMATCH")
        if got != want:
            sys.exit(1)

json.dump(report, open("_sca1close/tables/sow_w1w2_diff.json", "w"), indent=1)
print(f"\n{len(report)} rows targeted; diff written to _sca1close/tables/sow_w1w2_diff.json")
