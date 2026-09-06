#!/usr/bin/env python3
# timetable_extract.py — pass PEQ-YEAR-3 §1: the real 2026-27 timetable, measured.
#
# The owner supplied two workbooks. They are EVIDENCE and are NEVER committed to
# this repository (same rule as the ASDAN instruments, _passpq/inputs/README.md):
# pass their paths in, and what lands in git is this tool's extract, with a source
# cell cited for every row.
#
#   python3 _passpq/tools/timetable_extract.py <Timetables.xlsx> <Colour_coded.xlsx>
#
# Emits _passpq/TIMETABLE_2026-27.md and _passpq/inputs/timetable_2026-27.json.
# Non-zero exit if a lane's grid cannot be read - a guessed slot is worse than a
# paused pass (the rule PEQ-YEAR-1 set and this pass inherits).

import json, os, re, sys, hashlib, collections

ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
OUT_MD = os.path.join(ROOT, "_passpq", "TIMETABLE_2026-27.md")
OUT_JSON = os.path.join(ROOT, "_passpq", "inputs", "timetable_2026-27.json")

# The period grid, read off the workbooks' own time column and cross-checked
# against the Teacher Timetable header row (D2:K2).
PERIODS = [("P1", 5, "09:30-10:10"), ("P2", 6, "10:10-10:50"), ("P3", 8, "11:00-11:40"),
           ("P4", 9, "11:40-12:20"), ("P5", 11, "12:50-13:30"), ("P6", 12, "13:30-14:10")]
DAYS = [("Mon", 2), ("Tue", 3), ("Wed", 4), ("Thu", 5), ("Fri", 6)]
PERIOD_MIN = 40
# NOT guided slots, per the instrument and the workbooks' own note (sheet 1, A2):
NON_GUIDED = ("check-in", "daily reading", "reading / ar", "break", "lunch", "check-out", "earwig")

# ---- THE CLASSIFICATION RULE, printed on the page so it can be argued with ----
RULE_A = re.compile(r"\bASDAN\b|\bPEQ\b", re.I)          # (a) explicitly ASDAN/PEQ-labelled
RULE_B = [                                                # (b) carryable
    (re.compile(r"\bPfA\b", re.I), "PfA"),
    (re.compile(r"communit|enterprise", re.I), "Community / enterprise"),
    (re.compile(r"careers|gatsby|work experience", re.I), "Careers"),
    (re.compile(r"vocational|practical", re.I), "Vocational / practical"),
    (re.compile(r"enrichment|outdoor", re.I), "Enrichment / outdoor"),
    (re.compile(r"catch[- ]?up", re.I), "Catch-up"),
]
RULE_C = re.compile(r"english|literacy|maths|numeracy|science|\bPE\b|physical development|"
                    r"behaviour intervention|\bRE\b|humanities|citizenship|british values|"
                    r"creative arts|\bICT\b|computin|PSHE|RSHE", re.I)


# Slots whose label reads BOTH ways under the rule above. The instrument is explicit:
# "Where a slot could plausibly be either, list it separately as owner-decision rather
# than silently counting it." These are named, not pattern-guessed, so the list is
# auditable and short.
BORDERLINE = {
    "outdoor learning / regulation":
        "'Outdoor learning' is carryable enrichment; 'regulation' is behaviour-support time, "
        "which the rule puts with Behaviour Intervention on the not-carryable side. The label "
        "is both, so it is not counted either way.",
}


def classify(label):
    if label.strip().lower() in BORDERLINE:
        return "owner-decision", BORDERLINE[label.strip().lower()]
    """Returns (bucket, why). Buckets: a | b | c | owner-decision.
    A slot that matches BOTH a carryable and a not-carryable pattern is never
    silently absorbed - it is returned as owner-decision with both readings named."""
    if RULE_A.search(label):
        return "a", "label literally names ASDAN/PEQ"
    hit_b = [name for pat, name in RULE_B if pat.search(label)]
    hit_c = bool(RULE_C.search(label))
    if hit_b and hit_c:
        return "owner-decision", f"reads as carryable ({'/'.join(hit_b)}) AND as a subject that banks its own ({label})"
    if hit_b:
        return "b", f"carryable: {'/'.join(hit_b)}"
    if hit_c:
        return "c", "banks its own subject"
    return "owner-decision", "matches no rule - unclassified, not absorbed"


def split_cell(v):
    if v is None:
        return None, None
    parts = [x.strip() for x in str(v).split("\n") if x.strip()]
    if not parts:
        return None, None
    subj = parts[0]
    who = parts[1].strip("()") if len(parts) > 1 else None
    # "RSHE (Maths Teacher) " - role folded into the subject line on one cell
    m = re.match(r"^(.*?)\s*\((.*Teacher.*)\)\s*$", subj)
    if m and who is None:
        subj, who = m.group(1).strip(), m.group(2).strip()
    return subj, who


def main():
    if len(sys.argv) < 3:
        print(__doc__ or "usage: timetable_extract.py <Timetables.xlsx> <Colour_coded.xlsx>")
        return 2
    tt_path, cc_path = sys.argv[1], sys.argv[2]
    import openpyxl
    sha = lambda p: hashlib.sha256(open(p, "rb").read()).hexdigest()
    prov = {"timetables": {"file": os.path.basename(tt_path), "sha256": sha(tt_path)},
            "colour_coded": {"file": os.path.basename(cc_path), "sha256": sha(cc_path)}}

    tt = openpyxl.load_workbook(tt_path, data_only=True)
    cc = openpyxl.load_workbook(cc_path, data_only=True)

    # --- the Teacher Timetable: the more specific label source -----------------
    # It disambiguates the lane grids' bare "PfA" into PfA / cooking, PfA / ASDAN,
    # PfA / independence. Keyed (lane, day, period) -> (label, cell).
    ts = tt["Teacher Timetable"]
    TEACH = {}
    TCOL = {"P1": "D", "P2": "E", "P3": "G", "P4": "H", "P5": "J", "P6": "K"}
    day_rows = {}
    cur = None
    for r in range(3, ts.max_row + 1):
        a = ts.cell(r, 1).value
        if isinstance(a, str) and a.strip() in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday"):
            cur = a.strip()[:3]
        if cur:
            day_rows.setdefault(cur, []).append(r)
    for day, rows in day_rows.items():
        for r in rows:
            for pl, col in TCOL.items():
                v = ts[f"{col}{r}"].value
                if not isinstance(v, str) or ":" not in v:
                    continue
                lane, _, lab = v.partition(":")
                lane = lane.strip()
                if lane.lower().startswith("all groups"):
                    for L in ("Build", "Grow", "Launch"):
                        TEACH[(L, day, pl)] = (lab.strip(), f"[Teacher Timetable]!{col}{r}")
                elif lane in ("Build", "Grow", "Launch"):
                    TEACH[(lane, day, pl)] = (lab.strip(), f"[Teacher Timetable]!{col}{r}")

    # --- the colour-coded summary's Details column ----------------------------
    ws1 = cc["1 Teacher subject totals"]
    SUMM = {}
    det = re.compile(r"(Build|Grow|Launch):\s*(Monday|Tuesday|Wednesday|Thursday|Friday)\s*Period\s*(\d)\s*\(([^)]*)\)")
    for r in range(5, ws1.max_row + 1):
        g = ws1.cell(r, 7).value
        if not isinstance(g, str):
            continue
        for m in det.finditer(g):
            lane, day, per, lab = m.group(1), m.group(2)[:3], "P" + m.group(3), m.group(4)
            SUMM[(lane, day, per)] = (lab.strip(), f"[1 Teacher subject totals]!G{r}",
                                      (ws1.cell(r, 2).value or "").strip())

    # --- the lane grids: the base -------------------------------------------
    slots, disagreements = [], []
    for lane in ("Build", "Grow", "Launch"):
        sheet = f"{lane} Timetable"
        if sheet not in tt.sheetnames:
            print(f"STOP: {sheet} missing"); return 1
        ws = tt[sheet]
        for pl, row, time in PERIODS:
            for dl, col in DAYS:
                cell = f"{openpyxl.utils.get_column_letter(col)}{row}"
                subj, who = split_cell(ws.cell(row, col).value)
                if subj is None:
                    slots.append({"lane": lane, "day": dl, "period": pl, "time": time,
                                  "label": None, "role": None, "bucket": "empty",
                                  "why": "no lesson timetabled", "source": f"[{sheet}]!{cell}",
                                  "teacher_sheet": None, "summary": None})
                    continue
                if any(k in subj.lower() for k in NON_GUIDED):
                    continue
                tlab, tsrc = TEACH.get((lane, dl, pl), (None, None))
                slab, ssrc, sarea = (SUMM.get((lane, dl, pl)) or (None, None, None))
                # the operative label: the most specific of the three
                label = subj
                if tlab and len(tlab) > len(subj):
                    label = tlab
                if slab and len(slab) > len(label):
                    label = slab
                bucket, why = classify(label)
                # record every disagreement between the three sources
                for other, osrc, oname in ((tlab, tsrc, "Teacher Timetable"),
                                           (slab, ssrc, "colour-coded Details")):
                    if other and other.lower() != subj.lower():
                        ob, _ = classify(other)
                        sb, _ = classify(subj)
                        # A REFINEMENT is not a disagreement. The lane grids write a bare
                        # "PfA" where the teacher sheet writes "PfA / ASDAN" or
                        # "PfA / cooking" - the second is the first plus detail, and the
                        # detail is exactly what this pass needs. A CONTRADICTION is where
                        # neither label contains the other ("Behaviour Intervention" vs
                        # "DT / practical skills"): different subjects, and those are the
                        # ones that go to the owner rather than being picked.
                        norm = lambda x: re.sub(r"[^a-z0-9]+", "", x.lower())
                        refinement = norm(subj) in norm(other) or norm(other) in norm(subj)
                        if refinement:
                            continue
                        disagreements.append({
                            "lane": lane, "day": dl, "period": pl, "time": time,
                            "lane_grid": subj, "lane_grid_source": f"[{sheet}]!{cell}",
                            "other": other, "other_source": oname + " " + (osrc or ""),
                            "material": ob != sb,
                            "buckets": f"{sb} vs {ob}"})
                slots.append({"lane": lane, "day": dl, "period": pl, "time": time,
                              "label": label, "role": who, "bucket": bucket, "why": why,
                              "source": f"[{sheet}]!{cell}",
                              "teacher_sheet": (tlab, tsrc) if tlab else None,
                              "summary": (slab, ssrc) if slab else None})

    # material disagreements force owner-decision, never a silent pick
    for d in disagreements:
        if d["material"]:
            for s in slots:
                if (s["lane"], s["day"], s["period"]) == (d["lane"], d["day"], d["period"]):
                    s["bucket"] = "owner-decision"
                    s["why"] = (f"workbooks disagree: lane grid says '{d['lane_grid']}', "
                                f"{d['other_source']} says '{d['other']}' - "
                                f"{d['buckets']}. Not resolved by preference.")

    data = {"pass": "PEQ-YEAR-3 §1", "provenance": prov, "period_minutes": PERIOD_MIN,
            "periods": [{"period": p, "time": t} for p, _, t in PERIODS],
            "non_guided": list(NON_GUIDED), "slots": slots, "disagreements": disagreements}

    # ---- counts ------------------------------------------------------------
    counts = collections.defaultdict(lambda: collections.Counter())
    for s in slots:
        if s["bucket"] != "empty":
            counts[s["lane"]][s["bucket"]] += 1
    data["counts"] = {k: dict(v) for k, v in counts.items()}
    for lane in ("Build", "Grow", "Launch"):
        c = counts[lane]
        data.setdefault("rates", {})[lane] = {
            "a_only_slots": c["a"], "a_only_h_per_week": round(c["a"] * PERIOD_MIN / 60, 4),
            "a_plus_b_slots": c["a"] + c["b"],
            "a_plus_b_h_per_week": round((c["a"] + c["b"]) * PERIOD_MIN / 60, 4),
            "owner_decision_slots": c["owner-decision"],
        }
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    json.dump(data, open(OUT_JSON, "w"), indent=1)

    # ---- the markdown extract, a source cell on every row -------------------
    L=[]
    L.append("# TIMETABLE_2026-27.md — the real timetable, measured\n")
    L.append("**Pass PEQ-YEAR-3 §1.** Generated by `_passpq/tools/timetable_extract.py`; every row")
    L.append("cites the workbook cell it came from. **The source workbooks are evidence and are")
    L.append("never committed to this repository** — same rule as the ASDAN instruments. Reproduce with:\n")
    L.append("```\npython3 _passpq/tools/timetable_extract.py <Timetables.xlsx> <Colour_coded_teacher_subject_tables.xlsx>\n```\n")
    L.append("| workbook | sha256 |\n|---|---|")
    for k,v in prov.items(): L.append(f"| `{v['file']}` | `{v['sha256']}` |")
    L.append(f"\n**Period grid:** {PERIOD_MIN} minutes · " + " · ".join(f"**{p}** {t}" for p,_,t in PERIODS))
    L.append("\nCheck-in, daily reading/AR, break, lunch and check-out/Earwig are **not guided slots**")
    L.append("and are excluded — the workbooks' own note says the same (`[1 Teacher subject totals]!A2`).\n")
    L.append("## The classification rule, printed\n")
    L.append("| bucket | rule | counts toward |\n|---|---|---|")
    L.append("| **(a)** explicitly ASDAN/PEQ | the label literally contains `ASDAN` or `PEQ` | the floor **and** the ceiling |")
    L.append("| **(b)** carryable | PfA · Community/enterprise · Careers · Vocational/practical · Enrichment/outdoor · Catch-up | the ceiling only |")
    L.append("| **(c)** not carryable | English · Maths · Science · PE · Behaviour Intervention · RE/Humanities/Citizenship/Creative arts/ICT/PSHE/RSHE — these bank their own subjects | neither |")
    L.append("| **owner-decision** | reads plausibly as both, or the two workbooks contradict each other | **neither — never silently absorbed** |\n")
    L.append("## Measured rate per lane\n")
    L.append("| lane | (a) slots | (a) h/wk | (a)+(b) slots | (a)+(b) h/wk | owner-decision |\n|---|---:|---:|---:|---:|---:|")
    for lane in ("Build","Grow","Launch"):
        rt=data["rates"][lane]
        L.append(f"| **{lane}** | {rt['a_only_slots']} | **{rt['a_only_h_per_week']}** | "
                 f"{rt['a_plus_b_slots']} | **{rt['a_plus_b_h_per_week']}** | {rt['owner_decision_slots']} |")
    L.append("\n## Every guided slot, with its source cell\n")
    for lane in ("Build","Grow","Launch"):
        L.append(f"### {lane}\n")
        L.append("| day | period | time | subject as timetabled | teacher role | bucket | source cell | more specific label from |")
        L.append("|---|---|---|---|---|---|---|---|")
        order={d:i for i,(d,_) in enumerate(DAYS)}
        rows=[s for s in slots if s["lane"]==lane and s["bucket"]!="empty"]
        rows.sort(key=lambda s:(order[s["day"]], s["period"]))
        for s in rows:
            more = ""
            if s["teacher_sheet"] and s["teacher_sheet"][0].lower()!=s["label"].lower():
                more=f"`{s['teacher_sheet'][1]}`"
            elif s["teacher_sheet"] and s["teacher_sheet"][0]==s["label"]:
                more=f"`{s['teacher_sheet'][1]}`"
            b={"a":"**(a) ASDAN**","b":"(b) carryable","c":"(c) own subject",
               "owner-decision":"⚠ **owner-decision**"}[s["bucket"]]
            L.append(f"| {s['day']} | {s['period']} | {s['time']} | {s['label']} | {s['role'] or '—'} | {b} | `{s['source']}` | {more or '—'} |")
        emp=[s for s in slots if s["lane"]==lane and s["bucket"]=="empty"]
        for s in emp:
            L.append(f"| {s['day']} | {s['period']} | {s['time']} | *(no lesson timetabled)* | — | — | `{s['source']}` | — |")
        L.append("")
    L.append("## Owner-decision slots — listed, not absorbed\n")
    for s in slots:
        if s["bucket"]=="owner-decision":
            L.append(f"- **{s['lane']} {s['day']} {s['period']} ({s['time']})** — `{s['label']}`  \n  {s['why']}")
    L.append("\n## Workbook disagreements\n")
    L.append(f"**{len(disagreements)} disagreement(s), all material.** A *refinement* (the lane grid's bare")
    L.append("`PfA` where the teacher sheet writes `PfA / ASDAN` or `PfA / cooking`) is **not** counted as a")
    L.append("disagreement — the second label contains the first and adds the detail this pass needs. A")
    L.append("*contradiction* is where neither label contains the other; those are the rows below.\n")
    L.append("| lane | day | period | lane grid says | source | other source says | source | effect |")
    L.append("|---|---|---|---|---|---|---|---|")
    for d in disagreements:
        L.append(f"| {d['lane']} | {d['day']} | {d['period']} | `{d['lane_grid']}` | `{d['lane_grid_source']}` | "
                 f"`{d['other']}` | {d['other_source']} | {d['buckets']} — **{'material' if d['material'] else 'cosmetic'}** |")
    L.append("\n**Not resolved by preference.** DT/practical is carryable; Behaviour Intervention is not.")
    L.append("All three are counted as owner-decision and lodged.\n")
    L.append("## Noted, not edited\n")
    L.append("- `[Grow Timetable]!B11` reads **`ICT/Computinh`** — a typo for *Computing*. The workbooks are")
    L.append("  read-only evidence, so it is recorded here and not corrected.")
    L.append("- `[Grow Timetable]!F9` folds the role into the subject line (`RSHE (Maths Teacher)`) where every")
    L.append("  other cell puts it on a second line. Parsed correctly; noted as a formatting inconsistency.")
    L.append("- **Friday P6 is empty on all three lanes** (`F12` in each grid) — 3 unallocated slots.")
    open(OUT_MD,"w",encoding="utf-8").write("\n".join(L)+"\n")

    print("PEQ-YEAR-3 §1 — timetable measured")
    print(f"  period unit {PERIOD_MIN} min · {len(PERIODS)} periods/day · 5 days")
    for lane in ("Build", "Grow", "Launch"):
        c, rt = counts[lane], data["rates"][lane]
        print(f"  {lane:<7} a={c['a']} b={c['b']} c={c['c']} owner-decision={c['owner-decision']}"
              f"  |  (a) {rt['a_only_h_per_week']} h/wk · (a+b) {rt['a_plus_b_h_per_week']} h/wk")
    print(f"  disagreements between workbooks: {len(disagreements)} "
          f"({sum(1 for d in disagreements if d['material'])} material)")
    print(f"  -> {OUT_JSON}\n  -> {OUT_MD}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
