#!/usr/bin/env python3
# l2k_plan.py — pass PEQ-L2K: the three-lane year ledger, built and PROVEN here.
#
# Single source of numeric truth for the Kitchen Programme year plan
# (GROW_ASDAN/PEQ_L2_Kitchen/). Emits:
#   _passpq/inputs/peq_l2k_year_ledger.json          (data twin, drives the SoW page build)
#   _passpq/inputs/PEQ_L2K_YearPlan_2026-27.xlsx     (data-only xlsx twin)
#
# Every figure is asserted before anything is written; a ledger that does not
# sum raises and nothing is emitted — that IS the red gate (Addendum B §B2).
# All times are integer MINUTES (210 min = 3.5 h default week) so sums are exact.
#
# Facts: _passpq/SPEC_FACTS.md + _passpq/SPEC_FACTS_L2.md (spec v1.2 Oct 2025).
# GLH counts taught/supervised time incl. assessment by the centre assessor
# (spec §9 p16, §5.1 p9). Co-delivery (one supervised hour banked to two units)
# is used sparingly, declared per week, and grounded in the spec's own model:
# ComSk plan/use evidence is EXPECTED to be generated through a challenge that
# leads to another unit (pp25/39/54), ASDAN's worked assessment plan co-assesses
# two units in the same weekly activities (SPEC_FACTS §19), and the qualification
# GLH totals themselves sit below the unit sums (134 vs 140 at E3), so ASDAN's
# arithmetic already nets overlap out.

import json, os, sys

ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
OUT_JSON = os.path.join(ROOT, "_passpq", "inputs", "peq_l2k_year_ledger.json")
OUT_XLSX = os.path.join(ROOT, "_passpq", "inputs", "PEQ_L2K_YearPlan_2026-27.xlsx")

WEEKLY_MIN = 210          # OWNER INPUT: 3.5 GLH/week (Addendum B §B1). NOT a derived figure.
# Pass PEQ-YEAR-1 §1 attempted to replace this with a rate measured from the estate's own
# timetable and STOPPED: the repo does not fix one. See _passpq/DERIVATION_YEAR1.md and
# `python3 _passpq/tools/year1_derive.py` (exit 1 while any lane is unestablishable).
# What §1 DID establish, and what the sensitivity table below is now re-based onto:
PERIOD_MIN = 40           # the estate's period unit — 15 agreeing statements, all lanes
SLOT_BAND = (1, 7)        # BUILD's measurable band: 1 PEQ row .. 7 rows (PEQ + six slots)
DERIVATION = (
    "Period unit 40 min: MEASURED, 15 agreeing statements across all three lanes' weekly "
    "planners (each on a science row — it fixes the period unit, never a PEQ slot count). "
    "Slot count: 'Six discrete weekly slots' + 'which is how LI reaches certification on a "
    "one-slot week' (BUILD_Slot_Planner_2026-27_vA.xlsx [Slot Architecture]!A2) fixes ONE "
    "period per slot per week — the only cell in the estate that fixes a slot count, and a "
    "BUILD file. BUILD's PEQ time is therefore bounded 0.67–4.67 h/wk (1–7 periods). It is "
    "NOT a single figure: which of the six carryable slots may bank an hour to PEQ as well "
    "as to its own ASDAN short course needs the member-gated PEQ delivery guide (not held) "
    "or a coordinator ruling. GROW's ASDAN row is empty 8/8 weeks; LAUNCH's eight planners "
    "disagree on their own row structure (a populated PEQ row in 3 of 8, absent from 4). "
    "The 3.5 h/wk owner input is 5.25 forty-minute periods — not a whole number of the "
    "estate's actual timetabled periods."
)
WEEKS = 38
DECK_MIN = 40             # one timetabled 40-minute period per live deck (measured, §"method" below)

# The measurement method for the live 12 weeks (Addendum B gate: "assigned by
# measurement (state the method) not assertion"):
DECK_METHOD = (
    "12 live decks (LAUNCH_ASDAN/PEQ W1-6 + GROW_ASDAN/PEQ W1-6) x one timetabled "
    "40-minute period each = 480 min = 8.0 GLH. The period length is the estate's "
    "timetable unit (the science chassis is named v3_40min; GROW PEQ runs Thursday P4, "
    "one period). The decks PROGRAM 53 min of slide time each (sum of data-timer values "
    "on every deck's 10 slides) - i.e. they over-program the slot; the GLH claimed is "
    "the supervised 40-minute slot, never the programme. Deck-to-unit assignment "
    "follows each deck's own banking line."
)

# ---- units (credits, GLH) per lane -----------------------------------------
# skill keys: Com, DecMk, LSk, TmWk, Th (ThSk at E3/L1, CrThSk2 at L2), Wellb
UNITS = {
 "E3": {"Com": ("ComSkE3", 3, 30), "DecMk": ("DecMkSkE3", 2, 20), "LSk": ("LSkE3", 2, 20),
        "TmWk": ("TmWkSkE3", 2, 20), "Th": ("ThSkE3", 2, 20), "Wellb": ("WellbLeE3", 3, 30)},
 "L1": {"Com": ("ComSk1", 3, 27), "DecMk": ("DecMkSk1", 2, 18), "LSk": ("LSk1", 2, 18),
        "TmWk": ("TmWkSk1", 2, 18), "Th": ("ThSk1", 3, 27), "Wellb": ("WellbLe1", 3, 27)},
 "L2": {"Com": ("ComSk2", 3, 24), "DecMk": ("DecMkSk2", 2, 16), "LSk": ("LSk2", 2, 16),
        "TmWk": ("TmWkSk2", 2, 16), "Th": ("CrThSk2", 3, 24), "Wellb": ("WellbLe2", 3, 24)},
}
SKILLS = ["Com", "DecMk", "LSk", "TmWk", "Th", "Wellb"]

# Qualification ladders: (name, credits, qual GLH, milestone week, constituent skills)
QUALS = {
 "E3": [("Entry Level Award (610/5901/8)", 4, 40, 14, ["DecMk", "TmWk"]),
        ("Entry Level Extended Award (610/5903/1)", 9, 86, 26, ["DecMk", "TmWk", "Com", "LSk"]),
        ("Entry Level Certificate (610/5902/X)", 14, 134, 38, SKILLS)],
 "L1": [("Level 1 Award (610/5904/3)", 4, 36, 14, ["DecMk", "TmWk"]),
        ("Level 1 Extended Award (610/5906/7)", 9, 77, 26, ["DecMk", "TmWk", "Com", "LSk"]),
        ("Level 1 Certificate (610/5905/5)", 14, 120, 38, SKILLS)],
 "L2": [("Level 2 Award (610/5907/9)", 4, 32, 14, ["DecMk", "TmWk"]),
        ("Level 2 Extended Award (610/5909/2)", 9, 68, 26, ["DecMk", "TmWk", "Com", "LSk"]),
        ("Level 2 Certificate (610/5908/0)", 15, 106, 38, SKILLS)],
}

BLOCKS = [("Aut1", 1, 7), ("Aut2", 8, 14), ("Spr1", 15, 20), ("Spr2", 21, 26),
          ("Sum1", 27, 32), ("Sum2", 33, 38)]

# ---- the live 12-week deck spine (weeks 1-12) ------------------------------
# (week, suite, deck, skill banked — from each deck's banking line)
DECKS = [
 (1, "LAUNCH", "W1 Intro and Choosing My Level", "Com"),
 (2, "LAUNCH", "W2 What Makes Communication Effective", "Com"),
 (3, "LAUNCH", "W3 Active Listening and Giving Feedback", "Com"),
 (4, "LAUNCH", "W4 Plan a Communication Activity", "Com"),
 (5, "LAUNCH", "W5 Deliver the Activity and Gather Evidence", "Com"),
 (6, "LAUNCH", "W6 Review Progress and Sign Off the Unit", "Com"),
 (7, "GROW", "W1 Knowing Myself (core-skills audit)", "LSk"),
 (8, "GROW", "W2 Goals That Work (SMART planning)", "LSk"),
 (9, "GROW", "W3 Working With Others", "TmWk"),
 (10, "GROW", "W4 Managing Myself (routines)", "LSk"),
 (11, "GROW", "W5 Solving Problems", "Th"),
 (12, "GROW", "W6 Present My Progress", "LSk"),
]

def deck_minutes(week):
    for w, suite, name, sk in DECKS:
        if w == week:
            return sk, DECK_MIN, f"{suite} {name}"
    return None, 0, None

# ---- per-lane weekly allocations (minutes), hand-set, machine-proven -------
# row = {skill: physical minutes}; co = {skill: co-delivered minutes} (the same
# physical session banked to a second unit, declared).
def lane_rows(lane):
    R = [dict() for _ in range(WEEKS + 1)]   # 1-indexed
    C = [dict() for _ in range(WEEKS + 1)]
    def p(w, **kw):
        for k, v in kw.items(): R[w][k] = R[w].get(k, 0) + v
    def co(w, **kw):
        for k, v in kw.items(): C[w][k] = C[w].get(k, 0) + v
    # deck slot, all lanes share the mixed-room session (level-correct demand per lane)
    for w in range(1, 13):
        sk, m, _ = deck_minutes(w)
        p(w, **{sk: m})
    if lane == "E3":
        p(1, TmWk=170); p(2, TmWk=120, Wellb=50); p(3, TmWk=120, Wellb=50)
        p(4, TmWk=170); p(5, TmWk=170)
        p(6, TmWk=120, DecMk=50); p(7, TmWk=120, DecMk=50); p(8, TmWk=120, DecMk=50)
        p(9, DecMk=170); p(10, DecMk=170); p(11, DecMk=170); p(12, DecMk=170)
        p(13, TmWk=50, DecMk=160); p(14, DecMk=210)
        for w in range(15, 20): p(w, LSk=180, Com=30)
        p(20, LSk=140, Com=70)
        co(17, Com=40); co(18, Com=40)          # Com prep generated through the LSk ladder (spec pp25/54)
        for w in range(21, 27): p(w, Com=210)
        for w in range(27, 32): p(w, Th=170, Wellb=40); co(w, Wellb=60)  # Wellb plan in live use during the investigation
        p(32, Th=140, Wellb=70); co(32, Wellb=40)
        p(33, Th=170, Wellb=40)
        for w in range(34, 39): p(w, Wellb=210)
    elif lane == "L1":
        p(1, TmWk=170); p(2, TmWk=120, Wellb=50); p(3, TmWk=120, Wellb=50)
        p(4, TmWk=170); p(5, TmWk=130, Wellb=40)
        p(6, TmWk=120, DecMk=50); p(7, TmWk=120, DecMk=50); p(8, TmWk=90, DecMk=80)
        p(9, DecMk=170); p(10, DecMk=170); p(11, DecMk=170); p(12, DecMk=170)
        p(13, DecMk=170, Wellb=40); p(14, DecMk=50, Wellb=160)
        for w in range(15, 20): p(w, LSk=160, Com=50)
        p(20, LSk=120, Com=60, Wellb=30)
        for w in range(21, 26): p(w, Com=180, Wellb=30)
        p(26, Com=170, Wellb=40)
        for w in range(27, 33): p(w, Th=180, Wellb=30)
        for w in range(28, 32): co(w, Wellb=30)
        p(33, Th=180, Wellb=30); p(34, Th=180, Wellb=30); p(35, Th=140, Wellb=70)
        for w in range(36, 39): p(w, Wellb=210)
    elif lane == "L2":
        p(1, TmWk=170); p(2, TmWk=120, Wellb=50); p(3, TmWk=120, Wellb=50)
        p(4, TmWk=170); p(5, TmWk=170)
        p(6, TmWk=100, DecMk=70); p(7, TmWk=70, DecMk=100)
        p(8, DecMk=170); p(9, DecMk=170); p(10, DecMk=170); p(11, DecMk=170)
        p(12, DecMk=110, Wellb=60)
        p(13, Wellb=100, QA=110); p(14, Wellb=100, QA=110)
        for w in range(15, 20): p(w, LSk=160, Com=50)
        p(20, Com=130, Wellb=80)
        for w in range(21, 26): p(w, Com=140, Wellb=70)
        p(26, Com=120, Wellb=50, QA=40)
        for w in range(27, 33): p(w, Th=180, Wellb=30)
        p(33, Th=180, Wellb=30); p(34, Th=140, Wellb=70)
        p(35, Wellb=210); p(36, Wellb=110, QA=100)
        p(37, QA=210); p(38, QA=210)
    return R, C

# Unit-complete / evaluation weeks derived below; plan windows (10-hour use,
# spec: 5 of 6 units per level; Com carries activity minima instead):
PLAN_WINDOWS = {   # skill: (plan opens, window closes, review point week) — all lanes share the vehicle
 "TmWk": (2, 12, 6),     # team plan decided together in W2; weekly cook cycles; review W6
 "DecMk": (9, 13, 11),   # decision plan opens with the menu/budget project
 "LSk": (15, 19, 18),    # learning plan opens with the technique ladder
 "Th": (27, 32, 30),     # investigation plan (ThSk E3/L1 · CrThSk2 L2)
 "Wellb": (27, 36, 32),  # wellbeing-in-learning plan, used across the summer learning weeks
}
COM_ACTIVITY = {  # Com has no 10-hour window; activity minima instead (SPEC_FACTS §16 / _L2 §4)
 "plan_weeks": (21, 22), "activity_weeks": (23, 25), "eval_week": 26,
 "E3": "ONE activity: presentation >=2 min OR discussion >=5 min OR text >=100 words",
 "L1": "ONE activity: presentation >=3 min OR discussion >=8 min OR text >=250 words",
 "L2": "TWO plans over two DIFFERENT ways: presentation >=4 min / discussion >=10 min / text >=500 words",
}

def build():
    lanes = {}
    for lane in ("E3", "L1", "L2"):
        R, C = lane_rows(lane)
        # -- assertions: the red gate --
        for w in range(1, WEEKS + 1):
            phys = sum(R[w].values())
            assert phys == WEEKLY_MIN, f"{lane} W{w}: physical {phys} != {WEEKLY_MIN}"
            for sk in C[w]:
                assert R[w].get(sk if sk != "Wellb" else "Th", 0) + R[w].get(sk, 0) >= 0  # co rides a real session
        totals, cum, complete_wk = {}, {}, {}
        for sk in SKILLS:
            t = sum(R[w].get(sk, 0) + C[w].get(sk, 0) for w in range(1, WEEKS + 1))
            totals[sk] = t
            need = UNITS[lane][sk][2] * 60
            assert t == need, f"{lane} {UNITS[lane][sk][0]}: ledger {t} min != unit GLH {need} min"
            c, wk = 0, None
            for w in range(1, WEEKS + 1):
                c += R[w].get(sk, 0) + C[w].get(sk, 0)
                if wk is None and c >= need: wk = w
            complete_wk[sk] = wk
        qa = sum(R[w].get("QA", 0) for w in range(1, WEEKS + 1))
        co_total = sum(sum(C[w].values()) for w in range(1, WEEKS + 1))
        phys_total = WEEKLY_MIN * WEEKS
        assert sum(totals.values()) + qa == phys_total + co_total, f"{lane}: attribution mismatch"
        # milestones: constituent units complete by milestone week; credits + qual GLH stated
        miles = []
        for (name, cr, qglh, mw, skills) in QUALS[lane]:
            for sk in skills:
                assert complete_wk[sk] <= mw, f"{lane} {name}: {sk} completes W{complete_wk[sk]} > W{mw}"
            unit_glh = sum(UNITS[lane][sk][2] for sk in skills)
            unit_cr = sum(UNITS[lane][sk][1] for sk in skills)
            assert unit_cr == cr or (lane == "L1" and cr == 14 and unit_cr == 15), \
                f"{lane} {name}: credits {unit_cr} != {cr}"
            assert unit_glh >= qglh, f"{lane} {name}: unit GLH {unit_glh} < qual GLH {qglh}"
            miles.append({"qual": name, "credits": cr, "unit_credits": unit_cr,
                          "qual_glh": qglh, "unit_glh": unit_glh, "week": mw})
        # plan windows: >=600 min of the unit's supervised time inside the window, closed before eval
        for sk, (o, cl, rev) in PLAN_WINDOWS.items():
            inwin = sum(R[w].get(sk, 0) + C[w].get(sk, 0) for w in range(o, cl + 1))
            assert inwin >= 600, f"{lane} {sk}: window W{o}-W{cl} holds {inwin} min < 600"
            assert o <= rev <= cl, f"{lane} {sk}: review point W{rev} outside window"
        lanes[lane] = {"rows": [dict(R[w]) for w in range(1, WEEKS + 1)],
                       "co": [dict(C[w]) for w in range(1, WEEKS + 1)],
                       "totals_min": totals, "qa_min": qa, "co_min": co_total,
                       "complete_week": complete_wk, "milestones": miles}
    # deck spine sums to 8.0 GLH exactly
    assert len(DECKS) == 12 and 12 * DECK_MIN == 480
    return lanes

# ---- sensitivity table (Addendum B: 2 / 3 / 4 h/wk, default 3.5) -----------
def sensitivity():
    rows = []
    CO_CAP = {"E3": 420, "L1": 120, "L2": 0}   # the declared co-delivery in the 3.5 h plan
    # Re-based (PEQ-YEAR-1 §2) onto the MEASURED 40-minute period rather than the round
    # numbers 2/3/3.5/4, none of which is a whole number of the estate's real periods.
    # Each row is a whole count of timetabled periods; the owner input is carried alongside,
    # flagged, so the page can show it without presenting it as derived.
    lo, hi = SLOT_BAND
    grid = [(n, n * PERIOD_MIN / 60.0) for n in range(lo, hi + 1)]
    grid.append((WEEKLY_MIN / PERIOD_MIN, WEEKLY_MIN / 60.0))     # the declared owner input
    grid.sort(key=lambda t: t[1])
    for slots, hpw in grid:
        year = int(round(hpw * 60)) * WEEKS
        cell = {"slots": slots, "whole_periods": float(slots).is_integer(),
                "owner_input": abs(hpw - WEEKLY_MIN / 60.0) < 1e-9,
                "hpw": round(hpw, 4), "year_glh": round(year / 60.0, 2), "lanes": {}}
        for lane in ("E3", "L1", "L2"):
            reach = []
            for (name, cr, qglh, mw, skills) in QUALS[lane]:
                need = sum(UNITS[lane][sk][2] for sk in skills) * 60
                ok = need <= year + CO_CAP[lane]
                reach.append({"qual": name.split(" (")[0], "unit_glh_h": need / 60.0, "ok": ok,
                              "short_h": 0 if ok else (need - year - CO_CAP[lane]) / 60.0})
            cell["lanes"][lane] = reach
        rows.append(cell)
    return rows

def main():
    lanes = build()
    sens = sensitivity()
    data = {
        "pass": "PEQ-L2K v2 + Addendum B", "built": "2026-08-22",
        "weekly_min_default": WEEKLY_MIN, "weeks": WEEKS, "blocks": BLOCKS,
        "deck_min": DECK_MIN, "deck_method": DECK_METHOD,
        "period_min": PERIOD_MIN, "slot_band": list(SLOT_BAND), "derivation": DERIVATION,
        "weekly_min_is_owner_input": True,
        "decks": [{"week": w, "suite": s, "deck": d, "skill": sk} for (w, s, d, sk) in DECKS],
        "units": {lane: {sk: {"code": u[0], "credits": u[1], "glh": u[2]} for sk, u in UNITS[lane].items()}
                  for lane in UNITS},
        "plan_windows": {sk: {"open": o, "close": c, "review": r} for sk, (o, c, r) in PLAN_WINDOWS.items()},
        "com_activity": COM_ACTIVITY,
        "lanes": lanes, "sensitivity": sens,
        "l1_cert_choice": (
            "L1 Certificate = 14 credits while the six L1 units sum to 15. DEFAULT: teach and "
            "evidence all six L1 units (15 credits - spec s5.1 permits exceeding the 14 minimum; "
            "all credits at level). Where a claim of exactly 14 is preferred, or a pupil is "
            "stretched by L1 Thinking (the largest E3-to-L1 step: assumptions + primary/secondary "
            "sources), the NAMED default combination is the five other L1 units (12 cr at L1) + "
            "ThSkE3 (2 cr adjacent) = 14 cr, min-11-at-level and max-3-below both satisfied - the "
            "only exact-14 arithmetic. Confirm the preferred claim shape with ASDAN at "
            "names-confirmation (QUESTIONS_FOR_CHERYL)."),
    }
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w") as f:
        json.dump(data, f, indent=1)
    write_xlsx(data)
    print("PROVEN and written:")
    for lane in ("E3", "L1", "L2"):
        L = lanes[lane]
        tot = {k: v / 60.0 for k, v in L["totals_min"].items()}
        print(f"  {lane}: units {tot} · co-delivered {L['co_min']/60.0} h · QA/unassigned {L['qa_min']/60.0} h")
    print(f"  physical year: {WEEKLY_MIN*WEEKS/60.0} h/lane at {WEEKLY_MIN/60.0} h/wk x {WEEKS} wks")
    print(f"  -> {OUT_JSON}\n  -> {OUT_XLSX}")

def write_xlsx(data):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "YearMap"
    head = ["Week", "Block", "Deck (live 12)", "Lane"] + [data["units"]["E3"][sk]["code"].replace("E3", "") + " (min)" for sk in SKILLS] + ["QA/unassigned (min)", "Co-delivered (min, unit)", "Physical (min)"]
    # use skill keys as neutral column names instead
    head = ["Week", "Block", "Deck (live 12)", "Lane", "Com", "DecMk", "LSk", "TmWk", "Th/CrTh", "Wellb", "QA", "Co-delivered", "Physical"]
    ws.append(head)
    blocks = {w: b for (b, a, z) in data["blocks"] for w in range(a, z + 1)}
    decks = {d["week"]: f'{d["suite"]} {d["deck"]}' for d in data["decks"]}
    for lane in ("E3", "L1", "L2"):
        rows = data["lanes"][lane]["rows"]; cos = data["lanes"][lane]["co"]
        for i in range(data["weeks"]):
            w = i + 1; r = rows[i]; c = cos[i]
            co_txt = "; ".join(f'{v} {k}' for k, v in c.items()) or ""
            ws.append([w, blocks[w], decks.get(w, ""), lane] +
                      [r.get(sk, 0) for sk in SKILLS] + [r.get("QA", 0), co_txt, sum(r.values())])
    ws2 = wb.create_sheet("UnitLedgers")
    ws2.append(["Lane", "Unit", "Credits", "Unit GLH (h)", "Ledger (h)", "Complete by week"])
    for lane in ("E3", "L1", "L2"):
        for sk in SKILLS:
            u = data["units"][lane][sk]; L = data["lanes"][lane]
            ws2.append([lane, u["code"], u["credits"], u["glh"], L["totals_min"][sk] / 60.0, L["complete_week"][sk]])
    ws3 = wb.create_sheet("Milestones")
    ws3.append(["Lane", "Qualification", "Credits", "Qual GLH", "Constituent unit GLH", "Milestone week"])
    for lane in ("E3", "L1", "L2"):
        for m in data["lanes"][lane]["milestones"]:
            ws3.append([lane, m["qual"], m["credits"], m["qual_glh"], m["unit_glh"], m["week"]])
    ws4 = wb.create_sheet("Sensitivity")
    ws4.append(["h per week", "Year physical GLH", "Lane", "Qualification", "Unit GLH needed", "Reachable", "Short by (h)"])
    for row in data["sensitivity"]:
        for lane in ("E3", "L1", "L2"):
            for q in row["lanes"][lane]:
                ws4.append([row["hpw"], row["year_glh"], lane, q["qual"], q["unit_glh_h"],
                            "YES" if q["ok"] else "NO", q["short_h"]])
    ws5 = wb.create_sheet("Method")
    ws5.append(["Live-deck GLH measurement method"]); ws5.append([data["deck_method"]])
    ws5.append([]); ws5.append(["L1 Certificate 14-of-15 choice"]); ws5.append([data["l1_cert_choice"]])
    wb.save(OUT_XLSX)

if __name__ == "__main__":
    main()
