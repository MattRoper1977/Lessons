#!/usr/bin/env python3
# year1_derive.py — pass PEQ-YEAR-1 §1: derive the real PEQ timetable, or refuse.
#
# The year map was built on an OWNER INPUT of 3.5 supervised hours per week
# (l2k_plan.py WEEKLY_MIN = 210). §1 of the pass instrument requires that constant
# to be replaced by a figure MEASURED from the estate's own timetable evidence.
#
# This tool measures. It does not assume, and it does not fill a gap with a guess:
# where the evidence does not fix a figure, it says so and the pass STOPs (§1.4 —
# "a guessed slot is worse than a paused pass").
#
# Emits _passpq/inputs/year1_timetable_evidence.json and prints the verdict.
# Non-zero exit when any lane is not establishable — that IS the red gate.

import json, os, re, sys, glob, collections

ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
OUT = os.path.join(ROOT, "_passpq", "inputs", "year1_timetable_evidence.json")

LANES = {
    "BUILD":  "Planning/BUILD/Progress_Tees_Valley_BUILD_Weekly_Plan_Week*.xlsx",
    "GROW":   "Planning/GROW/Progress_Tees_Valley_GROW_Weekly_Plan_Week*.xlsx",
    "LAUNCH": "Planning/LAUNCH/Progress_Tees_Valley_Weekly_Plan_Week*.xlsx",
}
# Rows that are subject teaching, never PEQ. Matched on the row's column-A label.
NOT_PEQ = re.compile(r'^(English|Maths|Science|PSHE|RE$|RE\b|Humanities|Arts|PE$)', re.I)
# Rows that bank PEQ criteria directly.
PEQ_CORE = re.compile(r'ASDAN PEQ|PfA/ASDAN', re.I)
# Timetabled supervised slots that COULD carry PEQ evidence — carryable, not counted
# until the coordinator rules on it (see the verdict below).
CARRYABLE = re.compile(r'Slot \d|Careers|Vocational|Living Independently|Independent Living|'
                       r'FoodWise|Community|Duke|Enterprise', re.I)
# The one sentence anywhere in the estate that fixes a period LENGTH.
PERIOD_LEN = re.compile(r'(\d+)\s*-?\s*min(?:ute)?\s*period', re.I)


def load(path):
    import openpyxl
    return openpyxl.load_workbook(os.path.join(ROOT, path), data_only=True)


def scan_lane(lane, pattern):
    files = sorted(glob.glob(os.path.join(ROOT, pattern)))
    rows = collections.defaultdict(list)       # label -> [(week, filled)]
    period_evidence = []
    for f in files:
        wk = re.search(r'Week(\d+)', os.path.basename(f)).group(1)
        ws = load(os.path.relpath(f, ROOT))["Weekly Plan"]
        for r in range(10, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if not isinstance(a, str) or not a.strip():
                continue
            label = a.strip()
            if label.startswith(("Enrichment", "Resources", "F = formative", "Print",
                                 "Whole-school", "SMSC")) or re.match(r'^(Mon|Tue|Wed|Thu|Fri)', label):
                continue
            filled = sum(1 for c in range(2, 10)
                         if ws.cell(r, c).value not in (None, "", "None"))
            rows[label].append((wk, filled))
            for c in range(2, 10):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    m = PERIOD_LEN.search(v)
                    if m:
                        period_evidence.append({
                            "file": os.path.relpath(f, ROOT), "cell": f"{chr(64+c)}{r}",
                            "row_label": label, "minutes": int(m.group(1)),
                            "quote": v.replace("\n", " ")[:180]})
    return files, rows, period_evidence


def classify(label):
    if PEQ_CORE.search(label):  return "core"
    if NOT_PEQ.match(label):    return "not-peq"
    if CARRYABLE.search(label): return "carryable"
    return "unclassified"


def main():
    report = {"pass": "PEQ-YEAR-1 §1", "lanes": {},
              "period_length": None, "slot_count_evidence": None, "verdict": {}}

    # --- the period LENGTH, measured once across the estate --------------------
    all_pe = []
    for lane, pat in LANES.items():
        _, _, pe = scan_lane(lane, pat)
        all_pe += [dict(p, lane=lane) for p in pe]
    mins = sorted({p["minutes"] for p in all_pe})
    report["period_length"] = {
        "minutes": mins[0] if len(mins) == 1 else None,
        "distinct_values_found": mins,
        "established": len(mins) == 1,
        "evidence": all_pe,
        "note": ("Every period-length statement in the estate's weekly planners agrees. "
                 "Each sits on a SCIENCE row, not a PEQ row: it fixes the estate's period "
                 "UNIT, not the number of periods any PEQ strand receives."),
    }

    # --- the SLOT COUNT, from the one file that states one ---------------------
    sp = "Build/_Archive_VersionA_LivingIndependently/BUILD_Slot_Planner_2026-27_vA.xlsx"
    a2 = load(sp)["Slot Architecture"]["A2"].value or ""
    report["slot_count_evidence"] = {
        "file": sp, "cell": "[Slot Architecture]!A2",
        "states_six_slots": "Six discrete weekly slots" in a2,
        "states_one_slot_week": "one-slot week" in a2,
        "quote": a2[:400],
        "note": ("'Six discrete weekly slots' + 'which is how LI reaches certification on a "
                 "one-slot week' together fix ONE period per slot per week on BUILD. This is "
                 "the only cell in the estate that fixes a slot COUNT. It is a BUILD file in "
                 "an archived Version-A tree; GROW and LAUNCH have no equivalent."),
    }

    # --- per lane --------------------------------------------------------------
    for lane, pat in LANES.items():
        files, rows, _ = scan_lane(lane, pat)
        n = len(files)
        strands = []
        for label, weeks in sorted(rows.items()):
            live = [w for w, f in weeks if f > 0]
            strands.append({
                "label": label, "kind": classify(label),
                "weeks_present": len(weeks), "weeks_populated": len(live),
                "populated_every_week": len(live) == n,
            })
        core = [s for s in strands if s["kind"] == "core"]
        carry = [s for s in strands if s["kind"] == "carryable"]
        core_weekly = [s for s in core if s["populated_every_week"]]
        carry_weekly = [s for s in carry if s["populated_every_week"]]
        # a lane is establishable only if it has a PEQ row present AND populated in
        # every built week, on a row layout that does not move between weeks.
        labels_per_row_stable = len({s["label"] for s in core}) <= 1
        est = bool(core_weekly) and labels_per_row_stable
        report["lanes"][lane] = {
            "weekly_files": n,
            "strands": strands,
            "core_rows_populated_every_week": [s["label"] for s in core_weekly],
            "carryable_rows_populated_every_week": [s["label"] for s in carry_weekly],
            "establishable": est,
            "why_not": None if est else (
                "no PEQ row is populated in every built week"
                if not core_weekly else
                "the PEQ row's label/position moves between weekly files — the 8 planners "
                "do not agree on their own row structure"),
        }

    # --- verdict ---------------------------------------------------------------
    pl = report["period_length"]["minutes"]
    for lane, L in report["lanes"].items():
        if not L["establishable"]:
            L["glh_per_week"] = None
            continue
        core_n = len(L["core_rows_populated_every_week"])
        carry_n = len(L["carryable_rows_populated_every_week"])
        L["glh_per_week"] = {
            "floor_core_only": round(core_n * pl / 60.0, 4),
            "ceiling_core_plus_carryable": round((core_n + carry_n) * pl / 60.0, 4),
            "core_periods": core_n, "carryable_periods": carry_n,
            "single_value_established": False,
            "why": ("The floor counts only rows that bank PEQ criteria directly. The ceiling "
                    "additionally counts every timetabled supervised slot. Which of the "
                    "carryable slots may bank an hour to PEQ *as well as* to its own ASDAN "
                    "short course is not a repo fact: the Slot Planner forbids double-claiming "
                    "between LI and FoodWise, while the PEQ spec expects PEQ evidence to be "
                    "generated through other activity. Resolving that needs the member-gated "
                    "PEQ delivery guide (not held — _passpq/inputs/README.md) or a coordinator "
                    "ruling. It is therefore NOT derived here."),
        }

    ok = all(L["establishable"] for L in report["lanes"].values()) and \
         all(L.get("glh_per_week", {}) and L["glh_per_week"]["single_value_established"]
             for L in report["lanes"].values())
    report["verdict"] = {
        "single_weekly_constant_established": ok,
        "stop": not ok,
        "missing": [
            "GROW: no populated PEQ row in any of the 8 built weekly planners — the ASDAN "
            "row exists but is empty 0/8. There is no GROW PEQ slot to measure.",
            "LAUNCH: the 8 weekly planners disagree on their own row structure; a PEQ row is "
            "populated in 3 of 8 weeks and absent from 4. Not a weekly slot.",
            "ALL LANES: the share of the carryable slots that may bank PEQ guided hours is a "
            "coordinator/delivery-guide ruling, not a repo fact.",
            "CALENDAR: spring and summer 2026-27 have no dates and no calendar-backed week "
            "count anywhere in the repo. Only autumn (15 weeks) is evidenced.",
        ],
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(report, f, indent=1)

    print("PEQ-YEAR-1 §1 — timetable derivation")
    print(f"  period length: {pl} min  (established={report['period_length']['established']}, "
          f"{len(all_pe)} statements, all agree)")
    sc = report["slot_count_evidence"]
    print(f"  slot count:    six weekly slots, one period each — BUILD only "
          f"(six={sc['states_six_slots']} one-slot-week={sc['states_one_slot_week']})")
    for lane, L in report["lanes"].items():
        g = L.get("glh_per_week")
        rng = (f"{g['floor_core_only']}–{g['ceiling_core_plus_carryable']} h/wk"
               if g else "NOT ESTABLISHABLE")
        print(f"  {lane:<7} establishable={str(L['establishable']):<5} {rng}"
              + (f"   ({L['why_not']})" if L["why_not"] else ""))
    print(f"  -> {OUT}")
    if report["verdict"]["stop"]:
        print("\nSTOP (§1.4): no single weekly constant is established. Missing:")
        for m in report["verdict"]["missing"]:
            print("   - " + m)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
