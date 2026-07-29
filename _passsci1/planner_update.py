#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Planner update (§10.5): rewrite the Science row (row 12, 'Weekly Plan' sheet) of each weekly
plan for weeks 03-07, all three pathways, to match the science lessons actually built, and to
show the two-period design (BUILD/GROW) / three-lessons-per-week (LAUNCH).

UPDATE IN PLACE, never insert_rows — that sidesteps the openpyxl merged-range trap entirely
(insert_rows moves cell content but not merged-range metadata). All merged ranges in these
sheets are header/footer rows; row 12 carries none, so overwriting its cells is safe.

Read -> print the matched span -> update -> assert -> save (§7 tooling rules).
"""
import sys
import glob
import pathlib
import openpyxl

REPO = pathlib.Path("/workspace/lessons")
sys.path.insert(0, str(REPO / "_passsci1/inputs"))

BUILD = __import__("content_build")
GROW = __import__("content_grow")
LAUNCH = {m: __import__(m) for m in
          ["content_launch_t1", "content_launch_t2", "content_launch_t3",
           "content_launch_t4", "content_launch_t5"]}

# LAUNCH topic module per SoW week
LAUNCH_WEEK = {3: "content_launch_t1", 4: "content_launch_t2", 5: "content_launch_t3",
               6: "content_launch_t4", 7: "content_launch_t5"}

TWO_PERIOD = ("Runs across BOTH weekly 40-min periods: Period 1 → up to We Do 2; Period 2 = "
              "extended Independent Work → Lundy → Exit, where the practical and the Assessor "
              "Witness evidence are made. One print pack covers both periods.")


def clean(s):
    return " ".join(str(s).replace("—", "-").split())


def build_row(pathway, week):
    """Return dict of cell values for the Science row for one (pathway, week)."""
    if pathway == "BUILD":
        L = next(x for x in BUILD.LESSONS if x["week"] == week)
        mod = BUILD
    elif pathway == "GROW":
        L = next(x for x in GROW.LESSONS if x["week"] == week)
        mod = GROW
    else:  # LAUNCH — three lessons this week
        mod = LAUNCH[LAUNCH_WEEK[week]]
        Ls = mod.LESSONS
        arc = ["Discover", "Use", "Master"]
        outcome = " | ".join(f"{arc[i]}: {clean(Ls[i]['lo'])}" for i in range(3))
        vocab = ", ".join(dict.fromkeys(t for Ls_ in Ls for t, _ in Ls_["ko_vocab"]))
        files = "; ".join(f"Science_Teesside/Launch/{x['slug']}.html" for x in Ls)
        return {
            2: f"{clean(mod.SPINE)} - {clean(Ls[0]['sow']).split(' - ')[0]}",
            3: outcome,
            4: ("THREE lessons this week (Discover -> Use -> Master), one per 40-min period. "
                "Each a v5 whiteboard deck with 3-tier print pack + Assessor Witness Statement."),
            5: vocab[:250],
            6: files,
            7: "F",
            8: ("Tiered 3 ways (Supported/Standard/Stretch) at arrival, independent work and exit. "
                "Maths scaffolds at every tier. " + clean(mod.SENSORY)),
        }
    # BUILD / GROW single lesson, two periods
    vocab = ", ".join(t for t, _ in L["ko_vocab"])
    return {
        2: f"{clean(mod.SPINE)} - {clean(L['sow']).split(' - ')[0]}",
        3: clean(L["lo"]),
        4: (f"v5 whiteboard deck '{clean(L['title'])}': " + "; ".join(clean(g) for g in L["glance"])
            + ". " + TWO_PERIOD),
        5: vocab,
        6: f"Science_Teesside/{'Build' if pathway=='BUILD' else 'Grow'}/{L['slug']}.html "
           f"(3-tier print pack off the title slide)",
        7: "F",
        8: ("Tiered 3 ways (Supported/Standard/Stretch) at arrival, independent work and exit. "
            + clean(mod.SENSORY)),
    }


PATH_DIR = {"BUILD": "Planning/BUILD/*BUILD_Weekly_Plan_Week{:02d}*",
            "GROW": "Planning/GROW/*GROW_Weekly_Plan_Week{:02d}*",
            "LAUNCH": "Planning/LAUNCH/*Weekly_Plan_Week{:02d}*"}


def find_science_row(ws):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "science":
            return r
    return None


def main(dry=False):
    changed = 0
    for pathway in ("BUILD", "GROW", "LAUNCH"):
        for week in range(3, 8):
            hits = glob.glob(str(REPO / PATH_DIR[pathway].format(week)))
            assert len(hits) == 1, f"{pathway} W{week}: expected 1 workbook, found {hits}"
            wbpath = hits[0]
            wb = openpyxl.load_workbook(wbpath)
            ws = wb["Weekly Plan"]
            r = find_science_row(ws)
            assert r, f"no Science row in {wbpath}"
            row = build_row(pathway, week)
            before = {c: ws.cell(r, c).value for c in row}
            print(f"\n{pathway} W{week} (row {r}) BEFORE: {clean(before[3])[:70]!r}")
            print(f"                      AFTER : {clean(row[3])[:70]!r}")
            if not dry:
                for c, v in row.items():
                    ws.cell(r, c).value = v
                # post-condition: the outcome cell now equals what we set
                assert clean(ws.cell(r, 3).value) == clean(row[3]), "outcome not written"
                # also mirror the outcome into the Outcomes RAG sheet, science row, if present
                if "Outcomes RAG" in wb.sheetnames:
                    rag = wb["Outcomes RAG"]
                    for rr in range(1, rag.max_row + 1):
                        if str(rag.cell(rr, 3).value or "").strip().lower() == "science":
                            rag.cell(rr, 4).value = clean(row[3])[:200]
                            break
                wb.save(wbpath)
                changed += 1
    print(f"\n{'DRY RUN' if dry else 'UPDATED'} {changed} workbooks' Science rows (weeks 03-07 x 3 pathways)")


if __name__ == "__main__":
    main(dry="--dry" in sys.argv)
